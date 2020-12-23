__author__ = '{SunnyBhuee}'
__copyright__ = 'Copyright {2020}, {img2xl}'
__credits__ = ['{credit_list}']
__license__ = '{MIT}'
__version__ = '{1}.{0}.{0}'
__maintainer__ = '{SunnyBhuee}'
__email__ = '{23467467+SunnyBhuee@users.noreply.github.com}'
__status__ = '{Dev}'

import os

import PIL
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import utils


def img2xl(img_add, resize=True, shrink_pct=0.5):
    """
    Saves a user supplied image (.jpg or .png files) as an Excel file at the same location as original file.
    Final name of the file is same as the name of the image name supplied by user.
    @param img_add: str -> address of the image file with .png or .jpg extension at the end, supplied as raw string.
                           e.g. r'C:/Documents/temp.jpg'
    @param resize: bool -> The image will be resized if set to true (default) before exporting to excel file. If set
                            to false, the resulting excel file can potentially be large and process can take long time
                            to conver the image to Excel file.
                            e.g. True or False
    @param shrink_pct: float -> Activates only if resize is set to true. Helps shrink or expand an image by a factor.
                                The factor can be any number expressed as either int or float (default = 0.5)
                                Use a value lower than 1 if image needs to be shrunk before exporting to excel.
                                Use a value higher than 1 if image needs to be expanded before exporting to excel.

    @return: True if the process executes without errors.
    """
    img = img_import(img_add)

    if resize:
        img = resize_img(img, shrink_pct)

    excel_out = os.path.join(os.path.dirname(img_add), os.path.basename(img_add).split('.')[0] + '.xlsx')

    export_to_excel(img, excel_out)

    print('Process completed!')

    return True


def img_import(img_add):
    """
    Import Image using the Pillow Library.
    User supplies the image address as a raw string such as r'C:\raw.jpg'
    Returns imported image.
    If input is .png, it is converted to .jpg and image is returned.
    """
    print('Importing image using PIL.')
    supported = ['jpg', 'png']
    assert type(img_add) is str, "User input should be a string."
    img_fmt = img_add.split('.')[-1]
    assert img_fmt in supported, "Please check file extension. Only {} formats are supported".format(supported)

    img = PIL.Image.open(os.path.normpath(img_add))

    if not img_fmt == 'jpg':
        img = png2jpg(img)

    print('Image imported successfully!')
    return img


def png2jpg(img):
    """
    Converts .png image to .jpg format
    @param img: PIL.PngImagePlugin.PngImageFile
    @return: RGB format image file
    """
    print('Converting .png image to .jpg.')
    assert type(img) == PIL.PngImagePlugin.PngImageFile, "png2jpg function only works on .png files."
    jpeg_img = img.convert('RGB')
    print('Converted .png file to .jpg file.')
    return jpeg_img


def img_assertion(img):
    """
    Checks if image supplied is jpg, png, or PIL Image format.
    @param img: PIL.Image
    @return: True if all the assertions are met
    """
    assert type(img) == PIL.PngImagePlugin.PngImageFile \
           or type(img) == PIL.JpegImagePlugin.JpegImageFile \
           or type(img) == PIL.Image.Image, "Please ensure file format is either .jpg or .png"
    print('Image assertion check passed.')
    return True


def resize_img(img, shrink_pct=0.5):
    """
    Takes an image and resizes it according to user supplied shrink_pct (default 0.5).
    @param img: PIL.Image
    @param shrink_pct:  float -> Activates only if resize is set to true. Helps shrink or expand an image by a factor.
                                The factor can be any number expressed as either int or float (default = 0.5)
                                Use a value lower than 1 if image needs to be shrunk before exporting to excel.
                                Use a value higher than 1 if image needs to be expanded before exporting to excel.
    @return: resized PIL.Image
    """
    print('Resizing image by {}'.format(shrink_pct))
    img_assertion(img)
    b = int((float(img.size[0]) * float(shrink_pct)))
    h = int((float(img.size[1]) * float(shrink_pct)))
    img = img.resize((b, h), PIL.Image.ANTIALIAS)
    print('Resized image by {}'.format(shrink_pct))
    return img


def rgb2hex(r, g, b):
    """
    Takes (R,G,B) format and returns Hex color format.
    @param r: int
    @param g: int
    @param b: int
    @return: hex color format
    """
    return "{:02x}{:02x}{:02x}".format(r, g, b)


def export_to_excel(img, excel_out):
    """
    Takes an image imported using Pillow (PIL) library, and exports it into an Excel file sheet using openpyxl library.
    Cells of the Excel file are treated as Pixels and their color is based on the pixel color of the image.
    Original image's pixel colors are converted from RGB to Hex format before export.
    Columns of Excel sheet are also reset to a width of 3 as that ensures optimal output shape.
    @param img: PIL.Image
    @param excel_out: str -> address of the output excel file with .xlsx, supplied as raw string.
                           e.g. r'C:/Documents/temp.xlsx'
    @return: True if the process executes without errors.
    """
    print('Starting export of image to Excel file.')
    img_assertion(img)
    px = list(img.getdata())
    flat = [x for sets in px for x in sets]

    # open workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    count = 0
    for x in range(1, img.size[1] + 1):
        for y in range(1, img.size[0] + 1):
            cell_color = rgb2hex(flat[count], flat[count + 1], flat[count + 2])
            ws.cell(row=x, column=y).fill = PatternFill(start_color=cell_color,
                                                        end_color=cell_color,
                                                        fill_type='solid')
            ws.column_dimensions[utils.get_column_letter(y)].width = 3
            count += 3

    wb.save(excel_out)
    print('Image exported to Excel file saved at {}'.format(excel_out))

    return True
